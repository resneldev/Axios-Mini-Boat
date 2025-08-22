"""
Axios PC-Programm

Dieses Programm steuert einen ESP8266-basierten Sensor zur Messung der Wasserleitfähigkeit.
Es bietet sowohl manuelle Steuerung als auch automatischen Modus mit Echtzeit-Datenerfassung.

Funktionen:
- Echtzeit-Datenerfassung und Visualisierung
- Manuelle Steuerung 
- Automatischer Modus
- Datenexport nach Excel
- Grafische Benutzeroberfläche

Autoren: Resnel Ndemeze Sona, Chat GPT und Deepseek
Datum: 25.06.2025
Version: 1.0.0
"""

# ============ Bibliotheken ============
import asyncio  # für asynchrone Programmierung
import aiohttp  # für HTTP-Anfragen
import numpy as np  # für numerische Berechnungen
import matplotlib.pyplot as plt  # für Datenvisualisierung
from tkinter import *  # GUI mit Tkinter
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg  # Einbindung Matplotlib in Tkinter
import webbrowser  # zum Öffnen externer Webseiten
from openpyxl import Workbook, load_workbook  # für Excel-Dateien
import os  # für Datei- und Systemoperationen
import datetime  # für Zeitstempel

# Klasse zur Steuerung des ESP8266-Sensors
class ESP8266Controller:
    def __init__(self, ip_address):  # Konstruktor der Klasse, nimmt die IP-Adresse als Parameter
        self.base_url = f"http://{ip_address}/control"  # Basis-URL zur Steuerung des Sensors
        self.auto_mode_active = False  # Status des Automatikmodus
        self.address = ip_address  # IP-Adresse des Sensors speichern
        print(f"IP-Adresse: {ip_address}")  # Debug-Ausgabe

    async def send_command(self, action):  # Asynchrone Methode zum Senden eines Steuerbefehls an den Sensor
        print(f"[DEBUG] Sende Befehl: {action} an {self.address}")  # Debug-Ausgabe des gesendeten Befehls
        try:
            async with aiohttp.ClientSession() as session:  # Erstellt eine asynchrone HTTP-Session
                async with session.get(f"{self.base_url}?action={action}", timeout=1) as response:  # Führt GET-Anfrage aus
                    return await response.text()  # Gibt die Text-Antwort des Sensors zurück
        except Exception as e:  # Fehlerbehandlung
            return f"Verbindungsfehler: {e}"  # Fehlermeldung zurückgeben

    async def toggle_auto_mode(self, start):  # Aktiviert oder deaktiviert den Automatikmodus
        self.auto_mode_active = start  # Setzt den Modus entsprechend
        cmd = "auto_modus" if start else "manuell_modus"  # Wählt den passenden Befehl
        return await self.send_command(cmd)  # Führt den Befehl aus


class AxiosAPP:
    def __init__(self, controller):  # Konstruktor der Klasse, übernimmt den Controller als Parameter
        self.controller = controller  # Speichert den Controller zur späteren Verwendung
        self.bg_control_color = "#035E01"  # Hintergrundfarbe für Steuerungsbereich
        self.window = Tk()  # Erstellt das Hauptfenster
        self.window.title("Axios App")  # Setzt den Fenstertitel
        self.window.geometry("800x480")  # Setzt die Fenstergröße auf 800x480 Pixel
        self.window.iconbitmap("ship.ico")  # Setzt das Fenstersymbol
        self.window.config(bg=self.bg_control_color)  # Setzt die Hintergrundfarbe des Fensters

        self.max_data_points = 100  # Anzahl der anzuzeigenden Datenpunkte im Plot
        self.update_interval = 500  # Aktualisierungsintervall in Millisekunden
        self.auto_refresh = True  # Gibt an, ob der Plot automatisch aktualisiert wird
        self.ip_ready = False  # Status, ob IP-Adresse gültig eingegeben wurde

        self.x_data = np.linspace(0, 1, self.max_data_points)  # Initialisiert X-Achse mit gleichverteilten Werten
        self.y_data = np.zeros(self.max_data_points)  # Initialisiert Y-Achse mit Nullen

        self.excel_filename = "leitwert_data.xlsx"  # Dateiname für die Excel-Tabelle
        self.initialize_excel_file()  # Erstellt die Excel-Datei, falls sie noch nicht existiert
        self.after_id = None  # Platzhalter für ID des geplanten Updates (für späteres Abbrechen)
        self.y_lim = None
        
        # Erstellen der verschiedenen GUI-Rahmen zur Organisation
        self.control_frame = Frame(self.window, bg="#44C942", borderwidth=1, relief='groove')  # Steuerbereich rechts
        self.graph_frame = Frame(self.window, bg="#A3E7A3")  # Bereich für den Graphen links
        self.under_graph_frame = Frame(self.graph_frame, bg="#44C942")  # Unterer Bereich des Graphs
        self.under_graph_left = Frame(self.under_graph_frame, bg="#44C942")  # Linker Teil unten
        self.under_graph_right = Frame(self.under_graph_frame, bg='#18AF15')  # Rechter Teil unten
        self.IP_button_frame = Frame(self.under_graph_left, bg='#18AF15')  # Bereich für IP-Button
        self.mode_frame = Frame(self.control_frame, bg="#44C942")  # Modus-Schaltflächen (Auto/Manuell)
        self.manual_frame = Frame(self.control_frame, bg="#44C942")  # Bereich für manuelle Steuerung
        self.auto_frame = Frame(self.control_frame, bg="#44C942")  # Bereich für automatische Steuerung

        self.create_widgets()  # Erstellt alle grafischen Elemente (Buttons, Labels, etc.)

        # Layout der GUI-Rahmen mit pack()
        self.control_frame.pack(side='right', expand=YES, fill='both')  # Steuerbereich rechts ausfüllen
        self.graph_frame.pack(side='left', expand=YES, fill='both')  # Graphbereich links ausfüllen
        self.under_graph_frame.pack(side='bottom', expand=YES, fill='both')  # Unterer Bereich unten einfügen
        self.under_graph_left.pack(side=LEFT)  # Linker Teil des unteren Bereichs
        self.under_graph_right.pack(side=BOTTOM)  # Rechter Teil unten
        self.IP_button_frame.pack()  # IP-Button-Bereich anzeigen
        self.mode_frame.pack()  # Modusbereich anzeigen
        self.manual_frame.pack(side='bottom')  # Manuelle Steuerung anzeigen

        self.setup_plot()  # Initialisiert das Matplotlib-Diagramm
        self.window.after(self.update_interval, self.update_graph)  # Startet geplante Aktualisierung
        self.bind_keys()  # Verknüpft Tastenbefehle mit Aktionen
        self.window.protocol("WM_DELETE_WINDOW", self.on_closing)  # Definiert Verhalten beim Schließen des Fensters

        
    def create_widgets(self):  # Erstellt alle Steuerelemente der Benutzeroberfläche
        self.boutton_auto_mode()  # Button zum Aktivieren des Automatikmodus erstellen
        self.boutton_manuel_mode()  # Button zum Aktivieren des manuellen Modus erstellen
        self.pause_boutton()  # Button zum Pausieren/Fortsetzen der Graph-Aktualisierung erstellen
        self.Leitwert_value()  # Label zur Anzeige des aktuellen Leitwerts erstellen
        self.IP_Andern_Feld()  # Eingabefeld für die IP-Adresse erstellen
        self.IP_verbinden_boutton()  # Button zum Bestätigen der IP-Adresse erstellen
        self.setup_auto_controls()  # Steuerbuttons für Automatikmodus erstellen
        self.setup_manual_controls()  # Steuerbuttons für manuellen Modus erstellen
        self.speicher_boutton()  # Button zum Speichern des Graphs erstellen
        self.create_menu()  # Menüleiste erstellen
        self.statut()  # Statusanzeige oben im GUI anzeigen

    def boutton_auto_mode(self):  # Erstellt Button zum Umschalten in den Automatikmodus
        Button(self.mode_frame,  # Der Button wird im Modus-Frame platziert
               text="Auto",  # Beschriftung des Buttons
               font=("Helvetica", 20),  # Schriftart und -größe
               bg='#18AF15',  # Hintergrundfarbe
               fg='white',  # Textfarbe
               command=lambda: asyncio.create_task(self.activate_auto_modus()),  # Beim Klicken wird der Automatikmodus aktiviert
               width=6  # Breite des Buttons
        ).pack(side='left', padx=5)  # Platzierung links mit horizontalem Abstand

    def boutton_manuel_mode(self):  # Erstellt Button zum Umschalten in den manuellen Modus
        Button(self.mode_frame,
               text="Manuel",  # Beschriftung des Buttons
               font=("Helvetica", 20),
               bg='#18AF15',
               fg='white',
               command=lambda: asyncio.create_task(self.activate_manuell_modus()),  # Beim Klicken wird der manuelle Modus aktiviert
               width=6
        ).pack(side='right', padx=5)  # Platzierung rechts mit Abstand

    def pause_boutton(self):  # Erstellt Button zum Pausieren/Fortsetzen der Graph-Aktualisierung
        self.boutton_pause = Button(self.under_graph_right,  # Position im rechten unteren Bereich unter dem Graph
                                    text="Pause",  # Anfangstext des Buttons
                                    font=("Arial", 15),
                                    bg='#18AF15',
                                    fg='white',
                                    command=self.toggle_pause,  # Beim Klicken wird die automatische Aktualisierung ein-/ausgeschaltet
                                    width=10)
        self.boutton_pause.pack(side='bottom')  # Button unten im Bereich platzieren

    def speicher_boutton(self):  # Erstellt Button zum Speichern des aktuellen Graphen als Bilddatei
        Button(self.under_graph_right,
               text="Speichern",  # Beschriftung
               font=("Arial", 15),
               bg='#18AF15',
               fg='white',
               command=self.save_graph,  # Funktion zum Speichern wird aufgerufen
               width=10
        ).pack()  # Direkt einfügen (kein bestimmter Platz)

    def Leitwert_value(self):  # Erstellt ein Label zur Anzeige des aktuellen Leitwerts
        self.value_var = StringVar(value="Leitwert: 0.000 µS")  # Variable für dynamischen Textinhalt
        Label(self.under_graph_right,
              textvariable=self.value_var,  # Verknüpfung mit StringVar für Live-Anzeige
              font=('Arial', 12),
              bg='#18AF15'
        ).pack(side='top')  # Oben im Bereich anzeigen

    def IP_Andern_Feld(self):  # Erstellt Eingabefeld zur Eingabe der ESP-IP-Adresse
        self.IP_Feld = StringVar()  # Variable zur Speicherung des Texteingabewerts
        Label(self.under_graph_left,
              text="IP_ESP8266:",  # Beschriftung des Eingabefeldes
              font=('Arial', 8),
              bg='#18AF15'
        ).pack(side=LEFT)  # Beschriftung links vom Eingabefeld
        self.entry_ip = Entry(self.under_graph_left,
                              textvariable=self.IP_Feld,  # Verknüpftes Eingabefeld
                              width=10
        )
        self.entry_ip.pack(side=LEFT)  # Eingabefeld links platzieren

    def IP_verbinden_boutton(self):  # Erstellt Button zum Anwenden der eingegebenen IP-Adresse
        Button(self.IP_button_frame,
               text='verbinden',  # Beschriftung
               font=('Arial', 8),
               bg='#18AF15',
               fg='white',
               command=self.update_ip,  # Beim Klicken wird die IP aktualisiert
               width=10
        ).pack(side=RIGHT)  # Rechts in dem kleinen IP-Frame platzieren

    def statut(self):  # Erstellt eine Statusanzeige im Steuerbereich
        self.status_var = StringVar(value="Bereit - Nicht verbunden")  # Statusvariable mit Anfangszustand
        Label(self.control_frame,
              textvariable=self.status_var,  # Zeigt aktuellen Verbindungsstatus an
              font=('Arial', 10),
              bg="#69CF67"
        ).pack(side="top")  # Ganz oben im Steuerbereich platzieren

    def setup_auto_controls(self):  # Erstellt Steuerelemente für den Automatikmodus
        Label(self.auto_frame,  # Überschrift im Auto-Bereich
              text="Automatische Kontrolle",
              font=('Arial', 11),
              bg='#18AF15'
        ).pack(pady=5)  # Mit vertikalem Abstand einfügen

        self.start_btn = Button(self.auto_frame,  # Button zum Starten des Automatikmodus
                                text="Starten",
                                font=("Arial", 15),
                                bg='#18AF15',
                                command=lambda: asyncio.create_task(self.do_action("auto_start"))  # Startbefehl an ESP senden
        )
        self.start_btn.pack(pady=20)  # Mit Abstand platzieren

        self.stop_btn = Button(self.auto_frame,  # Button zum Stoppen des Automatikmodus
                               text="Stopp",
                               font=("Arial", 15),
                               bg='#18AF15',
                               command=lambda: asyncio.create_task(self.do_action("auto_stopp"))  # Stopbefehl an ESP senden
        )
        self.stop_btn.pack(pady=20)  # Mit Abstand platzieren

    def setup_manual_controls(self):  # Erstellt Steuerbuttons für manuellen Modus
        Button(self.manual_frame,  # Aufwärtsbewegung
               text="up",
               font=("Helvetica", 20),
               bg='#18AF15',
               fg='white',
               command=lambda: asyncio.create_task(self.do_action("GO_FORWARD_ON")),
               width=5
        ).grid(row=1, column=1, sticky=W)

        Button(self.manual_frame,  # Abwärtsbewegung
               text="down",
               font=("Helvetica", 20),
               bg='#18AF15',
               fg='white',
               command=lambda: asyncio.create_task(self.do_action("GO_BACKWARD_ON")),
               width=5
        ).grid(row=2, column=1, sticky=W)

        Button(self.manual_frame,  # Linksbewegung
               text="left",
               font=("Helvetica", 20),
               bg='#18AF15',
               fg='white',
               command=lambda: asyncio.create_task(self.do_action("GO_LEFT_ON")),
               width=5
        ).grid(row=2, column=0, sticky=W)

        Button(self.manual_frame,  # Rechtsbewegung
               text="right",
               font=("Helvetica", 20),
               bg='#18AF15',
               fg='white',
               command=lambda: asyncio.create_task(self.do_action("GO_RIGHT_ON")),
               width=5
        ).grid(row=2, column=2, sticky=W)

    async def activate_manuell_modus(self):  # Wechselt in den manuellen Modus
        await self.controller.toggle_auto_mode(False)  # Sendet Befehl zum Deaktivieren des Automatikmodus
        self.manual_frame.pack(side='bottom')  # Zeigt die manuellen Steuerungs-Buttons an
        self.auto_frame.pack_forget()  # Versteckt die Automatiksteuerung
        self.status_var.set("Manueller Modus aktiv")  # Aktualisiert den Status

    async def activate_auto_modus(self):  # Wechselt in den Automatikmodus
        await self.controller.toggle_auto_mode(True)  # Sendet Befehl zum Aktivieren des Automatikmodus
        self.auto_frame.pack(pady=10)  # Zeigt die Automatiksteuerung an
        self.manual_frame.pack_forget()  # Versteckt die manuellen Buttons
        self.status_var.set("Automatischer Modus aktiv")  # Aktualisiert den Status

    def setup_plot(self):  # Initialisiert das Diagramm zur Anzeige der Leitwertdaten
        self.fig, self.ax = plt.subplots(figsize=(7, 3), dpi=100)  # Erzeugt eine neue Matplotlib-Figur und Achse
        self.fig.patch.set_facecolor("#74bb80")  # Hintergrundfarbe des Diagrammfensters setzen
        self.ax.set_facecolor("#74bb80")  # Hintergrundfarbe des Diagrammbereichs setzen
        self.ax.set_title('Echtzeit-Leitwertanzeige in µS')  # Titel des Diagramms
        self.ax.set_xlabel('Zeit (s)')  # Beschriftung der X-Achse
        self.ax.set_ylabel('Leitwert (µS)')  # Beschriftung der Y-Achse
        self.ax.set_ylim(0, 200)  # Wertebereich der Y-Achse setzen
        self.ax.grid(True, color='black', linestyle='--', linewidth=0.5)  # Gitterlinien aktivieren
        self.line, = self.ax.plot(self.x_data, self.y_data, 'b-')  # Initiale Linie für die Messdaten
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.graph_frame)  # Bindet das Diagramm in das Tkinter-Fenster ein
        self.canvas.draw()  # Zeichnet das Diagramm
        self.canvas.get_tk_widget().pack(fill=BOTH, expand=True)  # Packt das Diagramm-Widget in das GUI

    def update_graph(self):  # Planmäßige Aktualisierung des Diagramms einrichten
        if self.auto_refresh:  # Nur wenn automatische Aktualisierung erlaubt ist
            asyncio.create_task(self._async_update_graph())  # Starte asynchrone Datenerfassung
        self.after_id = self.window.after(self.update_interval, self.update_graph)  # Wiederhole diese Methode regelmäßig

    def update_ip(self):  # Aktualisiert die IP-Adresse des Sensors
        new_ip = self.IP_Feld.get()  # Holt den eingegebenen IP-Wert
        if new_ip:
            self.controller.base_url = f"http://{new_ip}/control"  # Setzt neue URL für HTTP-Kommunikation
            self.controller.address = new_ip  # Aktualisiert die Adresse im Controller
            self.status_var.set(f"Verbunden mit {new_ip}")  # Setzt Statusanzeige auf verbunden
            self.ip_ready = True  # Markiert, dass IP gültig und bereit ist
            self.window.focus()  # Fokus zurück auf Hauptfenster
        else:
            self.status_var.set("Ungültige IP eingeben")  # Fehlermeldung bei leerem Feld
            self.ip_ready = False  # Kein gültiger Zustand

    async def _async_update_graph(self):  # Lädt neue Daten asynchron vom Sensor und aktualisiert das Diagramm sowie die Excel-Datei
        try:
            value_str = await self.controller.send_command("read_Leitwert")  # Fragt den aktuellen Leitwert vom ESP8266 ab
            #value = float(value_str) if value_str.isdigit() else 0  # Wandelt den Wert in eine Ganzzahl um, wenn gültig
            try:
                value = float(value_str)
                self.y_lim = value +5
                print(value)
            except ValueError:
                value = 0
            self.y_data = np.roll(self.y_data, -1)  # Verschiebt die Y-Daten um eins nach links
            self.y_data[-1] = value  # Fügt den neuen Wert am Ende des Arrays ein

            self.line.set_ydata(self.y_data)  # Aktualisiert die Y-Daten der geplotteten Linie
            self.ax.relim()  # Berechnet neue Achsengrenzen
            self.ax.autoscale_view(scalex=False)  # Skaliert nur die Y-Achse neu
            self.canvas.draw()  # Zeichnet das Diagramm neu
            self.value_var.set(f"Leitwert: {value} µS")  # Zeigt den aktuellen Wert im GUI an

            self.save_to_excel(value)  # Speichert den Wert mit Zeitstempel in der Excel-Datei
        except Exception as e:
            self.status_var.set(f"Fehler: {e}")  # Zeigt Fehlermeldung im Statusfeld

    def toggle_pause(self):  # Pausiert oder setzt die automatische Aktualisierung fort
        self.auto_refresh = not self.auto_refresh  # Umschalten zwischen aktiviert und pausiert
        self.boutton_pause.config(text="Fortsetzen" if not self.auto_refresh else "Pause")  # Beschriftung ändern
        status = "Grafik pausiert" if not self.auto_refresh else "Grafik aktiv"  # Statusnachricht anpassen
        self.status_var.set(status)  # Anzeige aktualisieren

    def initialize_excel_file(self):  # Erstellt die Excel-Datei, wenn sie noch nicht existiert
        if not os.path.exists(self.excel_filename):  # Prüft, ob Datei vorhanden ist
            wb = Workbook()  # Neues Arbeitsbuch erstellen
            ws = wb.active  # Aktives Arbeitsblatt auswählen
            ws.title = "Leitwert Data"  # Titel des Arbeitsblatts setzen
            ws.append(["Timestamp", "Leitwert"])  # Überschriftenzeile einfügen
            wb.save(self.excel_filename)  # Datei speichern

    def save_to_excel(self, value):  # Speichert aktuellen Leitwert in Excel-Tabelle
        try:
            wb = load_workbook(self.excel_filename)  # Öffnet bestehende Excel-Datei
            ws = wb.active  # Wählt aktives Arbeitsblatt
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Holt aktuellen Zeitstempel
            ws.append([timestamp, value])  # Fügt neue Zeile mit Zeit und Messwert ein
            wb.save(self.excel_filename)  # Speichert Datei
        except Exception as e:
            print(f"Fehler beim Excel-Speichern: {e}")  # Gibt Fehler in Konsole aus

    def bind_keys(self):  # Verknüpft Tasten mit Steuerbefehlen
        self.window.bind('<KeyPress-w>', self.key_press)  # W-Taste drücken: vorwärts
        self.window.bind('<KeyRelease-w>', self.key_release)  # W-Taste loslassen
        self.window.bind('<KeyPress-a>', self.key_press)  # A-Taste drücken: links
        self.window.bind('<KeyRelease-a>', self.key_release)  # A-Taste loslassen
        self.window.bind('<KeyPress-d>', self.key_press)  # D-Taste drücken: rechts
        self.window.bind('<KeyRelease-d>', self.key_release)  # D-Taste loslassen
        self.window.bind('<KeyPress-s>', self.key_press)  # S-Taste drücken: rückwärts
        self.window.bind('<KeyRelease-s>', self.key_release)  # S-Taste loslassen
        self.window.bind('n', lambda e: asyncio.create_task(self.activate_auto_modus()))  # Taste n: Auto-Modus aktivieren
        self.window.bind('m', lambda e: asyncio.create_task(self.activate_manuell_modus()))  # Taste m: Manuell-Modus aktivieren
        self.window.bind('<space>', lambda e: self.toggle_pause())  # Leertaste: Pause an/aus
        self.window.bind('p', lambda e: self.save_graph())  # Taste p: Graph speichern

    def create_menu(self):  # Erstellt Menüleiste mit Datei- und Hilfeoptionen
        menu_bar = Menu(self.window)  # Neues Menü erstellen
        file_menu = Menu(menu_bar, tearoff=0)  # Datei-Menü ohne Trennlinie
        file_menu.add_command(label="Manuell-Modus", command=lambda: asyncio.create_task(self.activate_manuell_modus()))  # Menüpunkt für Manuell
        file_menu.add_command(label="Auto-Modus", command=lambda: asyncio.create_task(self.activate_auto_modus()))  # Menüpunkt für Auto
        file_menu.add_command(label="Verlassen", command=self.on_closing)  # Menüpunkt zum Beenden
        menu_bar.add_cascade(label="Datei", menu=file_menu)  # Datei-Menü zur Menüleiste hinzufügen

        hilfe_menu = Menu(menu_bar, tearoff=0)  # Hilfe-Menü
        hilfe_menu.add_command(label="Über uns", command=lambda: webbrowser.open_new("https://drive.google.com/drive/folders/1bkagriOvjD9MTF8pxnf7xED7J-0VrQXM?usp=drive_link"))  # Öffnet Webseite
        hilfe_menu.add_command(label="Dokumentation", command=lambda: print("Dokumentation öffnen"))  # Platzhalter für Hilfe
        menu_bar.add_cascade(label="Hilfe", menu=hilfe_menu)  # Hilfe zur Menüleiste hinzufügen

        self.window.config(menu=menu_bar)  # Menüleiste dem Fenster zuweisen

    def on_closing(self):  # Wird beim Schließen des Fensters aufgerufen
        self.auto_refresh = False  # Stoppt die automatische Aktualisierung
        if self.after_id is not None:  # Wenn ein Update geplant ist
            try:
                self.window.after_cancel(self.after_id)  # Versucht das geplante Update abzubrechen
            except Exception:
                pass  # Fehler ignorieren
        plt.close(self.fig)  # Schliesst das Diagrammfenster
        if self.window:
            self.window.destroy()  # Zerstört das Tkinter-Fenster

    async def do_action(self, action):  # Führt eine Aktion durch, z.B. Bewegung oder Start/Stop
        await self.controller.send_command(action)  # Sendet den Befehl asynchron an den ESP8266

    def key_press(self, event):  # Wird ausgelöst, wenn eine Taste gedrückt wird
        if not self.ip_ready:  # Wenn keine IP-Adresse verbunden ist
            self.status_var.set("IP-Adresse noch nicht verbunden!")  # Hinweis anzeigen
            return
        if self.window.focus_get() == self.entry_ip:  # Wenn der Fokus auf dem Eingabefeld liegt
            return  # Keine Aktion ausführen
        keys = {"w": "GO_FORWARD_ON", "s": "GO_BACKWARD_ON", "a": "GO_LEFT_ON", "d": "GO_RIGHT_ON"}  # Tastenbelegung für Richtungen
        if not self.manual_frame.winfo_ismapped():  # Wenn nicht im manuellen Modus
            self.status_var.set("Tasten im Automatikmodus deaktiviert.")  # Hinweis anzeigen
            return
        if event.keysym in keys:  # Wenn gedrückte Taste bekannt ist
            asyncio.create_task(self.do_action(keys[event.keysym]))  # Aktion ausführen

    def key_release(self, event):  # Wird ausgelöst, wenn eine Taste losgelassen wird
        if not self.ip_ready or self.window.focus_get() == self.entry_ip:  # Wenn IP nicht bereit oder Eingabefeld aktiv
            return  # Nichts tun
        keys = {"w": "GO_FORWARD_OFF", "s": "GO_BACKWARD_OFF", "a": "GO_LEFT_OFF", "d": "GO_RIGHT_OFF"}  # Stop-Aktionen für Richtungen
        if event.keysym in keys:  # Wenn Taste bekannt ist
            asyncio.create_task(self.do_action(keys[event.keysym]))  # Stop-Aktion senden

    def save_graph(self):  # Speichert das aktuelle Diagramm als PNG-Datei
        try:
            self.fig.savefig("meine_grafik.png", dpi=300)  # Speichert mit hoher Auflösung
            self.status_var.set("Grafik erfolgreich gespeichert.")  # Erfolgsnachricht anzeigen
        except Exception as e:
            self.status_var.set(f"Fehler beim Speichern: {e}")  # Fehlermeldung anzeigen


def main():
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    loop = asyncio.new_event_loop()             # Erstelle einen neuen Event-Loop
    asyncio.set_event_loop(loop)                # Setze ihn als aktuellen Loop
    controller = ESP8266Controller("192.168.51.190")
    app = AxiosAPP(controller)

    async def tk_loop():
        while True:
            try:
                app.window.update()
                await asyncio.sleep(0.01)
            except TclError:
                break

    try:
        loop.run_until_complete(tk_loop())
    except KeyboardInterrupt:
        print("Programm manuell beendet.")

if __name__ == "__main__":  # Wird nur ausgeführt, wenn das Skript direkt gestartet wurde
    main()  # Ruft die Hauptfunktion auf
